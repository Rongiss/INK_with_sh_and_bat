from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def main():
    work_type = znach('C')
    work_ed = znach('E')
    print(len(work_type))
    print(len(work_ed))
    write_date(work_type=work_type, work_ed=work_ed)


def write_date(work_type: list, work_ed: list):
    wb = load_workbook('1.xlsx')
    ws = wb.active
    ws['A1'] = 12
    wb.save('2.xlsx')


def znach(colom: str) -> list:
    # заливаем файл
    wb = load_workbook('1.xlsx')

    # выбираем лист в таблице
    ws = wb['Справочник']

    # указываем № столбца
    max_row = ws.max_row
    max_col = ws.max_column
    l = []

    for row in range(1, max_row + 1):
        cell = ws['{}{}'.format(colom, row)]
        if cell.value != None:
            l.append(cell.value)
        else:
            break
    return l[1:]


if __name__ == '__main__':
    main()
